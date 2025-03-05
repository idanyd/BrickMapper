from pathlib import Path
import cv2
from skimage.metrics import structural_similarity as ssim
import re

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
from PIL import Image
import numpy as np



def create_comparison_excel(pieces_dir, renders_dir, results):
    """
    Create an Excel sheet showing matched pieces and their similarity scores
    using data from the results dictionary
    """

    def resize_image_for_excel(img_array, max_size=(100, 100)):
        """Resize image to fit in Excel cell"""
        img_pil = Image.fromarray(img_array)
        img_pil.thumbnail(max_size)
        img_bio = BytesIO()
        img_pil.save(img_bio, format="PNG")
        return img_bio.getvalue()

    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Piece Matches"

    # Set column headers
    ws["A1"] = "Element ID"
    ws["B1"] = "Page"
    ws["C1"] = "Step"
    ws["D1"] = "Piece Image"
    ws["E1"] = "Render Image"
    ws["F1"] = "Similarity Score"
    ws["G1"] = "Piece Filename"
    ws["H1"] = "Render Filename"

    # Set column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 40
    ws.column_dimensions["H"].width = 40

    # Start from row 2
    row = 2

    for element_id, pages in results.items():
        # Find the corresponding render image
        render_path = None
        for file in renders_dir.glob(f"{element_id}_*.*"):
            render_path = file
            break

        if render_path is None:
            continue

        render_img = cv2.imread(str(render_path))
        render_img = cv2.cvtColor(render_img, cv2.COLOR_BGR2RGB)

        for page_num, steps_data in pages.items():
            for step_info in steps_data:
                step_num = step_info["step"]
                similarity = step_info["similarity"]
                piece_num = step_info["piece"]

                # Find corresponding piece image
                piece_pattern = f"page_{page_num:03d}_step_{step_num:03d}_piece_{piece_num:03d}.*"
                piece_path = None
                for file in pieces_dir.glob(piece_pattern):
                    piece_path = file
                    break

                if piece_path is None:
                    continue

                piece_img = cv2.imread(str(piece_path))
                piece_img = cv2.cvtColor(piece_img, cv2.COLOR_BGR2RGB)

                # Add text data
                ws[f"A{row}"] = element_id
                ws[f"B{row}"] = page_num
                ws[f"C{row}"] = step_num
                ws[f"F{row}"] = f"{similarity:.3f}"
                ws[f"G{row}"] = piece_path.name
                ws[f"H{row}"] = render_path.name

                # Add images
                try:
                    # Add piece image
                    piece_data = resize_image_for_excel(piece_img)
                    piece_img_xl = XLImage(BytesIO(piece_data))
                    ws.row_dimensions[row].height = 75
                    ws.add_image(piece_img_xl, f"D{row}")

                    # Add render image
                    render_data = resize_image_for_excel(render_img)
                    render_img_xl = XLImage(BytesIO(render_data))
                    ws.add_image(render_img_xl, f"E{row}")

                except Exception as e:
                    print(f"Error adding images to row {row}: {e}")

                row += 1

    # Save the workbook
    output_path = "piece_comparisons.xlsx"
    wb.save(output_path)
    print(f"Created comparison Excel file: {output_path}")


def load_and_resize_image(image_bytes, target_size=(200, 200)):
    """Load image and resize it to a standard size using Pillow"""
    try:
        # Open the image with Pillow
        img = Image.open(BytesIO(image_bytes))

        # Resize the image
        img = img.resize(target_size, Image.Resampling.LANCZOS)

        # Convert to RGB if it's not already
        if img.mode != "RGB":
            img = img.convert("RGB")

        # Convert to numpy array for compatibility with other functions
        img_array = np.array(img)

        return img_array
    except Exception as e:
        print(f"Error loading image: {e}")
        return None


def get_element_id(filename):
    """Extract element ID from piece_renders filename"""
    match = re.match(r"(\d+)_\d+_\d+\.(jp(e?)g|png)", filename)
    if match:
        return match.group(1)
    return None


def get_page_step_info(filename):
    """Extract page and step numbers from pieces filename"""
    match = re.match(
        r"page_(\d+)_step_(\d+)_piece_(\d+)\.(jp(e?)g|png)", filename
    )
    if match:
        return int(match.group(1)), int(match.group(2)), int(match.group(3))
    return None, None


def compare_images(img1, img2):
    """Compare two images using SSIM"""
    score, _ = ssim(img1, img2, channel_axis=2, full=True)
    return score


def match_pieces(step_pieces, set_pieces):
    # Dictionary to store results
    results = (
        {}
    )  # {element_id: {page_num: [{step_num1, similarity}, {step_num2, similarity}, ...]}}

    # Resize each set piece image
    set_pieces = {
        element_id: [
            load_and_resize_image(render_img["image_data"])
            for render_img in render_imgs
        ]
        for element_id, render_imgs in set_pieces.items()
        if render_imgs[0]["image_data"] is not None
    }

    # Process each piece image
    for piece in step_pieces:
        piece_image, page_num, step_num, piece_num = (
            load_and_resize_image(piece["img"]),
            piece["page"],
            piece["step"],
            piece["piece"],
        )

        # Compare with each render image
        best_match = (None, -1)  # (element_id, similarity_score)
        for element_id, render_imgs in set_pieces.items():
            for render_img in render_imgs:
                # Iterate over each image matched for this element ID
                similarity = compare_images(piece_image, render_img)
                if similarity > best_match[1]:
                    best_match = (element_id, similarity)

        # If good match found (threshold may need to be adjusted)
        if best_match[1] > 0.6:
            element_id = best_match[0]
            if element_id not in results:
                results[element_id] = {}
            if page_num not in results[element_id]:
                results[element_id][page_num] = []
            if step_num not in results[element_id][page_num]:
                results[element_id][page_num].append(
                    {
                        "step": step_num,
                        "piece": piece_num,
                        "similarity": best_match[1],
                    }
                )

    return results


def get_piece_data(piece_file):
    """Load piece image and extract page, step, and piece numbers"""
    piece_img = cv2.imread(str(piece_file))
    if piece_img is None:
        return None
    page_num, step_num, piece_num = get_page_step_info(piece_file.name)
    if not (page_num and step_num):
        return None
    return {
        "img": piece_img,
        "page": page_num,
        "step": step_num,
        "piece": piece_num,
    }


def match_pieces_from_files(
    step_piece_renders_dir: Path, all_set_piece_renders_dir: Path
):
    # Load all set pieces images
    set_pieces = {}
    for render_file in all_set_piece_renders_dir.glob("*"):
        element_id = get_element_id(render_file.name)
        if element_id:
            img = cv2.imread(str(render_file))
            if img is not None:
                set_pieces[element_id] = [{"image_data": img}]

    # load all step pieces images
    step_pieces = [
        get_piece_data(piece_file)
        for piece_file in step_piece_renders_dir.glob("*")
    ]

    return match_pieces(step_pieces, set_pieces)


def main():
    pieces_dir = Path("data/processed_booklets/31147_2/pieces")
    renders_dir = Path("data/processed_booklets/31147_3/piece_renders")

    results = match_pieces_from_files(pieces_dir, renders_dir)

    create_comparison_excel(pieces_dir, renders_dir, results)
    """
    # Print results
    for element_id, pages in results.items():
        print(f"\nElement ID: {element_id}")
        for page, steps in pages.items():
            print(f"  Page {page}: Steps {steps}")
    """


if __name__ == "__main__":
    main()
